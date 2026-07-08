import io
import csv
from sqlalchemy.orm import Session
from openpyxl import Workbook
from typing import Dict, Any

from app.models.form import (
    Form,
    FormPublishedVersion,
    QuestionGroup,
    QuestionType,
)
from app.models.spatial import Basin, Wetland, Site, SpatialBoundary
from app.schemas.form import FormBlueprintResponse


class FormExportService:
    @staticmethod
    def generate_blueprint_json(
        db: Session, form_id: int, draft: bool = False
    ) -> Dict[str, Any]:
        db_form = db.query(Form).filter(Form.id == form_id).first()
        if not db_form:
            raise ValueError("Form not found")

        # Try to use published version unless draft=True is requested
        if not draft and db_form.active_version_id:
            pub_version = (
                db.query(FormPublishedVersion)
                .filter(FormPublishedVersion.id == db_form.active_version_id)
                .first()
            )
            if pub_version and pub_version.schema:
                # Return the pre-saved schema dict, but ensure the type
                # reflects the saved type in db
                schema_dict = dict(pub_version.schema)
                schema_dict["type"] = db_form.type
                return schema_dict

        # Fallback to live relational database tables
        active_groups = (
            db.query(QuestionGroup)
            .filter(
                QuestionGroup.form_id == form_id,
                QuestionGroup.deleted_at.is_(None),
            )
            .order_by(QuestionGroup.order.asc().nullslast())
            .all()
        )
        response_obj = FormBlueprintResponse.from_orm_model(
            db_form, active_groups
        )
        return response_obj.model_dump()

    @staticmethod
    def get_translation_label(
        item: Dict[str, Any], lang: str, default_val: str
    ) -> str:
        translations = item.get("translations") or []
        for trans in translations:
            t_lang = trans.get("lang") or trans.get("language")
            if t_lang == lang:
                val = (
                    trans.get("label")
                    or trans.get("name")
                    or trans.get("description")
                )
                if val:
                    return val
        if lang == "en":
            val = (
                item.get("label")
                or item.get("name")
                or item.get("description")
            )
            if val:
                return val
        return default_val

    # Maps canonical cascade question names to XLSForm filter expressions
    # that scope rows in spatial_cascade.csv to the correct hierarchy level.
    SPATIAL_CASCADE_FILTER_MAP = {
        "basin_id": "list_name='basin'",
        "wetland_id": ("list_name='wetland' and parent_key=${basin_id}"),
        "site_id": ("list_name='site' and parent_key=${wetland_id}"),
        "region_id": "list_name='region'",
        "district_id": ("list_name='district' and parent_key=${region_id}"),
        "subcounty_id": (
            "list_name='subcounty' and parent_key=${district_id}"
        ),
        "location_id": "list_name='subcounty'",
    }

    @classmethod
    def generate_xlsform(
        cls, db: Session, form_id: int, draft: bool = False
    ) -> io.BytesIO:
        blueprint = cls.generate_blueprint_json(db, form_id, draft)

        wb = Workbook()
        # openpyxl starts with a default active sheet, rename it to 'survey'
        survey_sheet = wb.active
        survey_sheet.title = "survey"
        choices_sheet = wb.create_sheet(title="choices")
        settings_sheet = wb.create_sheet(title="settings")

        # Map language codes to Kobo standard label syntax
        language_map = {
            "en": "English (en)",
            "sw": "Swahili (sw)",
            "fr": "French (fr)",
            "es": "Spanish (es)",
        }

        # Resolve languages dynamically
        languages = blueprint.get("languages") or ["en"]
        default_lang_code = blueprint.get("defaultLanguage") or "en"
        default_lang_name = language_map.get(
            default_lang_code,
            f"{default_lang_code.upper()} ({default_lang_code})",
        )

        # 1. Write settings sheet
        settings_headers = [
            "form_title",
            "form_id",
            "version",
            "default_language",
        ]
        settings_sheet.append(settings_headers)
        settings_sheet.append(
            [
                blueprint.get("name", "Form"),
                blueprint.get("name", "Form").lower().replace(" ", "_"),
                str(blueprint.get("version", 1)),
                default_lang_name,
            ]
        )

        # 2. Write survey and choices sheets
        survey_headers = ["type", "name"]
        for lang in languages:
            lang_name = language_map.get(lang, f"{lang.upper()} ({lang})")
            survey_headers.append(f"label::{lang_name}")
        survey_headers.extend(
            [
                "required",
                "hint",
                "relevant",
                "filter",
                "parameters",
                "constraint",
            ]
        )
        for lang in languages:
            lang_name = language_map.get(lang, f"{lang.upper()} ({lang})")
            survey_headers.append(f"constraint_message::{lang_name}")
        survey_sheet.append(survey_headers)

        choices_headers = ["list_name", "name"]
        for lang in languages:
            lang_name = language_map.get(lang, f"{lang.upper()} ({lang})")
            choices_headers.append(f"label::{lang_name}")
        choices_sheet.append(choices_headers)

        question_groups = blueprint.get("question_group", [])
        for group in question_groups:
            # Write group start
            group_name = group.get("name")

            # Map group labels dynamically
            group_labels = []
            for lang in languages:
                lbl = cls.get_translation_label(
                    group,
                    lang,
                    group.get("label") or group.get("description", ""),
                )
                group_labels.append(lbl)

            row = ["begin_group", group_name]
            row.extend(group_labels)
            # required, hint, relevant, filter, parameters, constraint
            row.extend(["", "", "", "", "", ""])
            # constraint_message for each language
            row.extend(["" for _ in languages])
            survey_sheet.append(row)

            questions = group.get("question", [])
            for q in questions:
                q_name = q.get("name")
                q_type = q.get("type", "text")
                required_val = "yes" if q.get("required") else "no"
                hint_val = (
                    q.get("tooltip", {}).get("text", "")
                    if isinstance(q.get("tooltip"), dict)
                    else ""
                )

                # Map question labels dynamically
                q_labels = []
                for lang in languages:
                    lbl = cls.get_translation_label(
                        q, lang, q.get("label", "")
                    )
                    q_labels.append(lbl)

                # Dependencies mapping (relevancy rules)
                relevant_val = ""
                deps = q.get("dependency", [])
                if deps:
                    rules = []
                    for dep in deps:
                        dep_q = dep.get("question")
                        dep_val = dep.get("value")
                        if dep_q and dep_val is not None:
                            rules.append(f"${{{dep_q}}} = '{dep_val}'")
                    op = (
                        " and " if q.get("dependencyRule") == "AND" else " or "
                    )
                    relevant_val = op.join(rules)

                # Map type to Kobo standard
                xls_type = "text"
                is_cascade = False
                filter_val = ""

                if q_type == QuestionType.cascade.value:
                    xls_type = "select_one_from_file spatial_cascade.csv"
                    is_cascade = True
                    filter_val = cls.SPATIAL_CASCADE_FILTER_MAP.get(q_name, "")
                    if not filter_val:
                        q_api = q.get("api") or {}
                        cascade_level = q_api.get("list_name") or q_api.get(
                            "initial"
                        )
                        if cascade_level:
                            filter_val = f"list_name='{cascade_level}'"
                elif q_type in ("integer", QuestionType.number.value):
                    xls_type = "integer" if q_type == "integer" else "decimal"
                elif q_type == QuestionType.text.value:
                    xls_type = "text"
                elif q_type == QuestionType.date.value:
                    xls_type = "date"
                elif q_type in (QuestionType.image.value, "photo"):
                    xls_type = "image"
                elif q_type in (QuestionType.geo.value, "geopoint", "gps"):
                    xls_type = "geopoint"
                elif q_type in (
                    QuestionType.option.value,
                    QuestionType.multiple_option.value,
                    "select_one",
                    "select_multiple",
                ):
                    # Check if it uses external file for cascade
                    q_extra = q.get("extra") or {}
                    if q_extra.get("cascade") or "cascade" in q_name.lower():
                        xls_type = "select_one_from_file spatial_cascade.csv"
                        is_cascade = True
                        filter_val = cls.SPATIAL_CASCADE_FILTER_MAP.get(
                            q_name, ""
                        )
                    else:
                        prefix_val = (
                            "select_multiple"
                            if q_type
                            in (
                                QuestionType.multiple_option.value,
                                "select_multiple",
                            )
                            else "select_one"
                        )
                        suffix = ""
                        if (
                            q.get("allowOther")
                            or q.get("allow_other")
                            or q_extra.get("allowOther")
                            or q_extra.get("allow_other")
                        ):
                            suffix = " or_other"
                        xls_type = f"{prefix_val} option_{q_name}{suffix}"

                param_val = ""
                if is_cascade:
                    param_val = "value=name label=label"

                constraint_val = ""
                constraint_messages = []
                rule_val = q.get("rule")
                if rule_val and isinstance(rule_val, dict):
                    min_val = rule_val.get("min")
                    max_val = rule_val.get("max")
                    if min_val is not None and max_val is not None:
                        constraint_val = f". >= {min_val} and . <= {max_val}"
                    elif min_val is not None:
                        constraint_val = f". >= {min_val}"
                    elif max_val is not None:
                        constraint_val = f". <= {max_val}"

                for lang in languages:
                    msg = ""
                    if constraint_val and rule_val:
                        min_val = rule_val.get("min")
                        max_val = rule_val.get("max")
                        if lang == "sw":
                            if min_val is not None and max_val is not None:
                                msg = (
                                    f"Thamani lazima iwe kati ya "
                                    f"{min_val} na {max_val}"
                                )
                            elif min_val is not None:
                                msg = (
                                    f"Thamani lazima iwe kubwa kuliko "
                                    f"au sawa na {min_val}"
                                )
                            elif max_val is not None:
                                msg = (
                                    f"Thamani lazima iwe ndogo kuliko "
                                    f"au sawa na {max_val}"
                                )
                        elif lang == "fr":
                            if min_val is not None and max_val is not None:
                                msg = (
                                    f"La valeur doit être comprise entre "
                                    f"{min_val} et {max_val}"
                                )
                            elif min_val is not None:
                                msg = (
                                    f"La valeur doit être supérieure "
                                    f"ou égale à {min_val}"
                                )
                            elif max_val is not None:
                                msg = (
                                    f"La valeur doit être inférieure "
                                    f"ou égale à {max_val}"
                                )
                        elif lang == "es":
                            if min_val is not None and max_val is not None:
                                msg = (
                                    f"El valor debe estar entre "
                                    f"{min_val} y {max_val}"
                                )
                            elif min_val is not None:
                                msg = (
                                    f"El valor debe ser mayor "
                                    f"o igual a {min_val}"
                                )
                            elif max_val is not None:
                                msg = (
                                    f"El valor debe ser menor "
                                    f"o igual a {max_val}"
                                )
                        else:  # en or other default
                            if min_val is not None and max_val is not None:
                                msg = (
                                    f"Value must be between "
                                    f"{min_val} and {max_val}"
                                )
                            elif min_val is not None:
                                msg = (
                                    f"Value must be greater than "
                                    f"or equal to {min_val}"
                                )
                            elif max_val is not None:
                                msg = (
                                    f"Value must be less than "
                                    f"or equal to {max_val}"
                                )
                    constraint_messages.append(msg)

                q_row = [xls_type, q_name]
                q_row.extend(q_labels)
                q_row.extend(
                    [
                        required_val,
                        hint_val,
                        relevant_val,
                        filter_val,
                        param_val,
                        constraint_val,
                    ]
                )
                q_row.extend(constraint_messages)
                survey_sheet.append(q_row)

                # Populate choices if multiple choice and not external cascade
                if (
                    q_type
                    in (
                        QuestionType.option.value,
                        QuestionType.multiple_option.value,
                        "select_one",
                        "select_multiple",
                    )
                    and not is_cascade
                ):
                    options = q.get("option") or []
                    if isinstance(options, str):
                        options = []
                    for opt in options:
                        opt_name = opt.get("value") or opt.get("name")

                        opt_labels = []
                        for lang in languages:
                            lbl = cls.get_translation_label(
                                opt,
                                lang,
                                opt.get("label") or opt.get("name") or "",
                            )
                            opt_labels.append(lbl)

                        opt_row = [f"option_{q_name}", opt_name]
                        opt_row.extend(opt_labels)
                        choices_sheet.append(opt_row)

            # Write group end
            end_row = ["end_group", ""]
            end_row.extend(["" for _ in languages])
            # required, hint, relevant, filter, parameters, constraint
            end_row.extend(["", "", "", "", "", ""])
            # constraint_message for each language
            end_row.extend(["" for _ in languages])
            survey_sheet.append(end_row)

        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        return file_stream

    @staticmethod
    def generate_spatial_cascade_csv(db: Session) -> io.StringIO:
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(["list_name", "name", "label", "parent_key"])

        # 1. Export Basins
        basins = db.query(Basin).order_by(Basin.name).all()
        for b in basins:
            writer.writerow(["basin", str(b.id), b.name, ""])

        # 2. Export Spatial Boundaries
        # (Regions, Districts/Counties, Sub-counties)
        boundaries = (
            db.query(SpatialBoundary)
            .order_by(SpatialBoundary.level, SpatialBoundary.name)
            .all()
        )
        for s in boundaries:
            list_name = "subcounty"
            if s.level == 1:
                list_name = "region"
            elif s.level == 2:
                list_name = "district"

            parent_key = str(s.parent_id) if s.parent_id else ""
            writer.writerow([list_name, str(s.id), s.name, parent_key])

        # 3. Export Wetlands
        wetlands = db.query(Wetland).order_by(Wetland.name).all()
        for w in wetlands:
            writer.writerow(["wetland", str(w.id), w.name, str(w.basin_id)])

        # 4. Export Sites
        sites = db.query(Site).order_by(Site.name).all()
        for s in sites:
            writer.writerow(["site", str(s.id), s.name, str(s.wetland_id)])

        output.seek(0)
        return output
