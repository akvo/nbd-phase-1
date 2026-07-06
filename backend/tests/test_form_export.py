from fastapi.testclient import TestClient
from app.main import app
from app.models.user import User
from app.config.auth import JWT_SECRET, JWT_ALGORITHM
import jwt

client = TestClient(app)


def get_auth_headers(
    db_session, email="admin_form_test@nbd.org", role="Admin"
):
    user = db_session.query(User).filter(User.email == email).first()
    if not user:
        user = User(email=email, role=role, is_active=True)
        db_session.add(user)
        db_session.commit()
    token = jwt.encode({"email": email}, JWT_SECRET, algorithm=JWT_ALGORITHM)
    return {"Authorization": f"Bearer {token}"}


def test_export_json_endpoint(db_session):
    headers = get_auth_headers(db_session)
    # Create a dummy form
    form_res = client.post(
        "/api/v1/forms",
        json={"name": "Export Test Form", "type": 1},
        headers=headers,
    )
    assert form_res.status_code == 201
    form_id = form_res.json()["id"]

    # Try exporting
    response = client.get(
        f"/api/v1/forms/{form_id}/export/json", headers=headers
    )
    assert response.status_code == 200
    res_data = response.json()
    assert res_data["name"] == "Export Test Form"


def test_export_xlsform_endpoint(db_session):
    headers = get_auth_headers(db_session)
    form_res = client.post(
        "/api/v1/forms",
        json={"name": "XLSForm Test", "type": 1},
        headers=headers,
    )
    assert form_res.status_code == 201
    form_id = form_res.json()["id"]

    response = client.get(
        f"/api/v1/forms/{form_id}/export/xlsform", headers=headers
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == (
        "application/vnd.openpyxlformats-officedocument.spreadsheetml.sheet"
    )


def test_export_cascade_csv_endpoint(db_session):
    headers = get_auth_headers(db_session)
    response = client.get(
        "/api/v1/reference/spatial-cascade/csv", headers=headers
    )
    assert response.status_code == 200
    assert response.headers["content-type"] == "text/csv; charset=utf-8"


def test_export_xlsform_with_cascade(db_session):
    headers = get_auth_headers(db_session)
    # 1. Create a form
    form_res = client.post(
        "/api/v1/forms",
        json={"name": "XLSForm Cascade Test", "type": 1},
        headers=headers,
    )
    assert form_res.status_code == 201
    form_id = form_res.json()["id"]

    # 2. Create group
    group_res = client.post(
        "/api/v1/question-groups",
        json={
            "form_id": form_id,
            "name": "grp_spatial",
            "label": "Spatial Group",
        },
        headers=headers,
    )
    assert group_res.status_code == 201
    group_id = group_res.json()["id"]

    # 3. Create a cascade type question and
    # an option type question named basin_id
    q_res = client.post(
        "/api/v1/questions",
        json={
            "form_id": form_id,
            "question_group_id": group_id,
            "name": "custom_wetland_question",
            "label": "Select Wetland",
            "type": "cascade",
            "order": 1,
        },
        headers=headers,
    )
    assert q_res.status_code == 201

    q_opt_res = client.post(
        "/api/v1/questions",
        json={
            "form_id": form_id,
            "question_group_id": group_id,
            "name": "basin_id",
            "label": "Select Basin",
            "type": "option",
            "order": 2,
        },
        headers=headers,
    )
    assert q_opt_res.status_code == 201

    # 4. Export to XLSForm
    response = client.get(
        f"/api/v1/forms/{form_id}/export/xlsform", headers=headers
    )
    assert response.status_code == 200

    # 5. Inspect the generated spreadsheet structure using openpyxl
    import io
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    survey_sheet = wb["survey"]

    # Find the row for custom_wetland_question and basin_id
    rows = list(survey_sheet.iter_rows(values_only=True))
    header = rows[0]
    type_idx = header.index("type")
    name_idx = header.index("name")

    q_row = None
    q_opt_row = None
    for r in rows[1:]:
        if r[name_idx] == "custom_wetland_question":
            q_row = r
        elif r[name_idx] == "basin_id":
            q_opt_row = r

    assert q_row is not None
    assert q_row[type_idx] == "select_one_from_file spatial_cascade.csv"

    assert q_opt_row is not None
    assert q_opt_row[type_idx] == "select_one option_basin_id"


def test_export_xlsform_with_allow_other(db_session):
    headers = get_auth_headers(db_session)
    # 1. Create a form
    form_res = client.post(
        "/api/v1/forms",
        json={"name": "XLSForm Allow Other Test", "type": 1},
        headers=headers,
    )
    assert form_res.status_code == 201
    form_id = form_res.json()["id"]

    # 2. Create group
    group_res = client.post(
        "/api/v1/question-groups",
        json={
            "form_id": form_id,
            "name": "grp_options",
            "label": "Options Group",
        },
        headers=headers,
    )
    assert group_res.status_code == 201
    group_id = group_res.json()["id"]

    # 3. Create a question with allow_other = True
    q_res = client.post(
        "/api/v1/questions",
        json={
            "form_id": form_id,
            "question_group_id": group_id,
            "name": "preferred_color",
            "label": "Preferred Color",
            "type": "option",
            "order": 1,
            "extra": {"allowOther": True},
        },
        headers=headers,
    )
    assert q_res.status_code == 201

    # 4. Export to XLSForm
    response = client.get(
        f"/api/v1/forms/{form_id}/export/xlsform", headers=headers
    )
    assert response.status_code == 200

    # 5. Inspect the generated spreadsheet structure using openpyxl
    import io
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    survey_sheet = wb["survey"]

    rows = list(survey_sheet.iter_rows(values_only=True))
    header = rows[0]
    type_idx = header.index("type")
    name_idx = header.index("name")

    q_row = None
    for r in rows[1:]:
        if r[name_idx] == "preferred_color":
            q_row = r

    assert q_row is not None
    assert q_row[type_idx] == "select_one option_preferred_color or_other"


def test_export_xlsform_with_validation_rules(db_session):
    headers = get_auth_headers(db_session)
    # 1. Create a form
    form_res = client.post(
        "/api/v1/forms",
        json={"name": "XLSForm Rules Test", "type": 1},
        headers=headers,
    )
    assert form_res.status_code == 201
    form_id = form_res.json()["id"]

    # Set languages on the form in db
    from app.models.form import Form

    db_form = db_session.query(Form).filter(Form.id == form_id).first()
    db_form.languages = ["en", "fr"]
    db_session.commit()

    # 2. Create group
    group_res = client.post(
        "/api/v1/question-groups",
        json={
            "form_id": form_id,
            "name": "grp_rules",
            "label": "Rules Group",
        },
        headers=headers,
    )
    assert group_res.status_code == 201
    group_id = group_res.json()["id"]

    # 3. Create question with rule (min and max)
    q_res = client.post(
        "/api/v1/questions",
        json={
            "form_id": form_id,
            "question_group_id": group_id,
            "name": "water_ph",
            "label": "pH Level",
            "type": "number",
            "order": 1,
            "rule": {"min": 2, "max": 10, "allowDecimal": True},
        },
        headers=headers,
    )
    assert q_res.status_code == 201

    # 4. Export to XLSForm
    response = client.get(
        f"/api/v1/forms/{form_id}/export/xlsform", headers=headers
    )
    assert response.status_code == 200

    # 5. Inspect the generated spreadsheet structure using openpyxl
    import io
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    survey_sheet = wb["survey"]

    rows = list(survey_sheet.iter_rows(values_only=True))
    header = rows[0]

    assert "constraint" in header
    assert "constraint_message::English (en)" in header
    assert "constraint_message::French (fr)" in header

    name_idx = header.index("name")
    constraint_idx = header.index("constraint")
    msg_en_idx = header.index("constraint_message::English (en)")
    msg_fr_idx = header.index("constraint_message::French (fr)")

    q_row = None
    for r in rows[1:]:
        if r[name_idx] == "water_ph":
            q_row = r

    assert q_row is not None
    assert q_row[constraint_idx] == ". >= 2 and . <= 10"
    assert q_row[msg_en_idx] == "Value must be between 2 and 10"
    assert q_row[msg_fr_idx] == "La valeur doit être comprise entre 2 et 10"
